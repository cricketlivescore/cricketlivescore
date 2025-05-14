
from flask import Flask, render_template, send_file, jsonify
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

app = Flask(__name__)

MATCH_URL = "https://crex.com/scoreboard/U5P/1JK/66th-Match/11/10/sco-vs-uae-66th-match-mens-cwc-league-2-2023-27/live"

def fetch_match_data():
    try:
        res = requests.get(MATCH_URL)
        soup = BeautifulSoup(res.text, 'html.parser')
        team1 = soup.find_all("div", class_="team")[0].text.strip()
        team2 = soup.find_all("div", class_="team")[1].text.strip()
        score_tag = soup.find("div", class_="score")
        score = score_tag.text.strip() if score_tag else "Loading..."
        rr = "7.82"
        overs = "15.2"
        wickets = "4"
        ball_speed = "139.4 km/h"
        match_odds = {"PR": "64%", "DSG": "36%"}
        commentary = [
            "15.2 | 1 run",
            "15.1 | 4 runs - brilliant drive!",
            "14.6 | OUT! Caught at deep square",
            "14.5 | Dot ball",
            "14.4 | 6 runs - over long on!"
        ]
        batsmen = [
            {"name": "Joe Root", "runs": 45, "balls": 32, "fours": 4, "sixes": 1, "sr": 140.6},
            {"name": "Dinesh Karthik", "runs": 30, "balls": 20, "fours": 2, "sixes": 2, "sr": 150.0}
        ]
        bowlers = [
            {"name": "Mujeeb Ur Rahman", "overs": 3.0, "runs": 21, "wickets": 1, "eco": 7.0},
            {"name": "Bjorn Fortuin", "overs": 2.0, "runs": 16, "wickets": 0, "eco": 8.0}
        ]
        players = [
            {"name": "Joe Root", "country": "ENG", "age": 34, "flag": "https://cricketvectors.akamaized.net/Teams/S.png"},
            {"name": "Mujeeb Ur Rahman", "country": "AFG", "age": 23, "flag": "https://cricketvectors.akamaized.net/Teams/Y.png"},
            {"name": "Dinesh Karthik", "country": "IND", "age": 39, "flag": "https://cricketvectors.akamaized.net/Teams/0.png"},
        ]
        return {
            "team1": team1,
            "team2": team2,
            "score": score,
            "overs": overs,
            "rr": rr,
            "wickets": wickets,
            "ball_speed": ball_speed,
            "match_odds": match_odds,
            "commentary": commentary,
            "batsmen": batsmen,
            "bowlers": bowlers,
            "players": players
        }
    except Exception as e:
        return {"team1": "Error", "team2": "", "score": str(e), "players": [], "commentary": []}

@app.route('/')
def dashboard():
    data = fetch_match_data()
    return render_template("dashboard.html", data=data)

@app.route('/overlay')
def overlay():
    data = fetch_match_data()
    return render_template("overlay.html", data=data)

@app.route('/download')
def download_excel():
    data = fetch_match_data()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Playing XI"
    ws1.append(["Name", "Country", "Age"])
    for p in data["players"]:
        ws1.append([p["name"], p["country"], p["age"]])

    ws2 = wb.create_sheet("Batting")
    ws2.append(["Name", "Runs", "Balls", "4s", "6s", "SR"])
    for b in data["batsmen"]:
        ws2.append([b["name"], b["runs"], b["balls"], b["fours"], b["sixes"], b["sr"]])

    ws3 = wb.create_sheet("Bowling")
    ws3.append(["Name", "Overs", "Runs", "Wickets", "Economy"])
    for b in data["bowlers"]:
        ws3.append([b["name"], b["overs"], b["runs"], b["wickets"], b["eco"]])

    file_path = "match_data.xlsx"
    wb.save(file_path)
    return send_file(file_path, as_attachment=True)

# API endpoints
@app.route('/api/score')
def api_score():
    return jsonify(fetch_match_data())

@app.route('/api/commentary')
def api_commentary():
    return jsonify(fetch_match_data()["commentary"])

@app.route('/api/players')
def api_players():
    return jsonify(fetch_match_data()["players"])

if __name__ == '__main__':
    app.run(debug=True, port=5000)
